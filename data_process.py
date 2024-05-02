import openpyxl

def read_csv():
  """
  Reads data from a CSV file and stores it in separate lists for each column.
  
  Args:
    csv_file_path (str): The path to the CSV file.
  """
  workbook = openpyxl.load_workbook('Book1.xlsx')

  # Select the active sheet (you can also select a specific sheet by name: workbook['Sheet1'])
  sheet = workbook.active

  # Create empty lists to store data from each column
  columns = [[] for _ in range(sheet.max_column)]
  counter = 0

  # Iterate through each row and store data from each column
  for row in sheet.iter_rows(values_only=True):
    counter += 1
    for i, cell_value in enumerate(row):
      columns[i].append(cell_value)
  workbook.close()
  return columns
  # Close the workbook
  

