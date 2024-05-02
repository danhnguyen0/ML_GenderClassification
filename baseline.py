import openpyxl
from text_processing import readJSON, common_grams, bag, get_gram_id,remove_rows_with_none_elements
from features import feat_cnt, constant
from sklearn.linear_model import LinearRegression
from metrics import measure
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression



# Open the Excel file
excel_file_path = 'Book1.xlsx'
workbook = openpyxl.load_workbook(excel_file_path)

# Select the active sheet (you can also select a specific sheet by name: workbook['Sheet1'])
sheet = workbook.active

# Create empty lists to store data from each column
columns = [[] for _ in range(sheet.max_column)]
counter = 0
# Iterate through each row and store data from each column
for row in sheet.iter_rows(values_only=True):
    counter +=1
    for i, cell_value in enumerate(row):
        
        columns[i].append(cell_value)


# Close the workbook
workbook.close()

columns = remove_rows_with_none_elements(columns)

train_data = columns[0:2587]
test_data = columns[2587:]
counter = 0

train_texts = train_data[0]
top_unigrams = common_grams(train_texts, 1, 1000, no_stopwords=False)

uni_bag = bag(top_unigrams)
uni_id = get_gram_id(uni_bag)
X = [feat_cnt(t, uni_bag, uni_id, 1, no_stopwords=False) for t in train_texts]
X = constant([X])
y = [1 if d == 'M' else 0 for d in train_data[1]] 


print(X)
model = LinearRegression()
model.fit(X, y)
model_b = LinearRegression()
model_b.fit(X, y)

test_texts = test_data[0]
X_test = [feat_cnt(t, uni_bag, uni_id, 1, no_stopwords=False) for t in test_texts]
X_test = constant([X_test])
y_test = [1 if d == 'M' else 0 for d in test_data[1]]
y_pred = model.predict(X_test)
y_pred_b = model_b.predict(X_test) 
print("Unbalanced: (acc, TP, TN, FP, FN, BER) = {}".format(measure(y_test, y_pred, 5)))
print("Balanced: (acc, TP, TN, FP, FN, BER) = {}".format(measure(y_test, y_pred_b, 5))) 





""" path = "./australian_user_reviews.json.gz"
data = []
for d in readJSON(path):
    data.append(d)

 # Num reviews = 59305, roughly 80/20 train/test split
reviews = []
for user in data:
    for r in user["reviews"]:
        reviews.append(r)
train_data = reviews[0:47500]
test_data = reviews[47500:]

# Baseline model
# Baseline uses unigrams, remove punctuation (keeps stopwords + no stemming)
train_texts = [d["review"] for d in train_data]
top_unigrams = common_grams(train_texts, 1, 1000, no_stopwords=False)
uni_bag = bag(top_unigrams)
uni_id = get_gram_id(uni_bag)
X = [feat_cnt(t, uni_bag, uni_id, 1, no_stopwords=False) for t in train_texts]
X = constant([X])
y = [1 if d["recommend"] == True else 0 for d in train_data]
print(X)
model = LogisticRegression(max_iter=1000)
model.fit(X, y)
model_b = LogisticRegression(max_iter=1000, class_weight="balanced")
model_b.fit(X, y)

test_texts = [d["review"] for d in test_data]
X_test = [feat_cnt(t, uni_bag, uni_id, 1, no_stopwords=False) for t in test_texts]
X_test = constant([X_test])
y_test = [1 if d["recommend"] == True else 0 for d in test_data]
y_pred = model.predict(X_test)
y_pred_b = model_b.predict(X_test)

print("Unbalanced: (acc, TP, TN, FP, FN, BER) = {}".format(measure(y_test, y_pred, 5)))
print("Balanced: (acc, TP, TN, FP, FN, BER) = {}".format(measure(y_test, y_pred_b, 5))) """